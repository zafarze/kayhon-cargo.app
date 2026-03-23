import io
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.permissions import IsAdminUser

from ..models import Package, ClientProfile


class ReportExportView(APIView):
    """
    Генерирует Excel-отчет по посылкам/клиентам за выбранный период.
    GET /api/reports/export/?type=packages&date_from=2026-01-01&date_to=2026-03-31
    Параметры type: packages | clients | finance
    """
    permission_classes = [IsAdminUser]

    def get(self, request):
        report_type = request.query_params.get('type', 'packages')
        date_from_str = request.query_params.get('date_from')
        date_to_str = request.query_params.get('date_to')

        # Пытаемся распарсить даты. Если не переданы — берём последние 30 дней
        try:
            date_from = datetime.strptime(date_from_str, '%Y-%m-%d') if date_from_str else datetime.now() - timedelta(days=30)
            date_to = datetime.strptime(date_to_str, '%Y-%m-%d') if date_to_str else datetime.now()
            date_to = date_to.replace(hour=23, minute=59, second=59)
        except ValueError:
            date_from = datetime.now() - timedelta(days=30)
            date_to = datetime.now()

        wb = openpyxl.Workbook()

        if report_type == 'clients':
            self._build_clients_sheet(wb, date_from, date_to)
        elif report_type == 'finance':
            self._build_finance_sheet(wb, date_from, date_to)
        else:
            # По умолчанию — отчет по посылкам
            self._build_packages_sheet(wb, date_from, date_to)

        # Сохраняем в память и отдаём файл
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        date_suffix = datetime.now().strftime('%Y-%m-%d')
        filename = f"kayhon_report_{report_type}_{date_suffix}.xlsx"

        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Access-Control-Expose-Headers'] = 'Content-Disposition'
        return response

    # ─────────────────────────────────────────
    #  ОТЧЁТ 1: ПОСЫЛКИ
    # ─────────────────────────────────────────
    def _build_packages_sheet(self, wb, date_from, date_to):
        ws = wb.active
        ws.title = "Посылки"

        # Стили
        header_fill = PatternFill("solid", fgColor="1A3C6E")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="D0D7E3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        # Заголовок листа
        ws.merge_cells('A1:I1')
        title_cell = ws['A1']
        period_str = f"Период: {date_from.strftime('%d.%m.%Y')} – {date_to.strftime('%d.%m.%Y')}"
        title_cell.value = f"Отчет по посылкам — Kayhon Cargo | {period_str}"
        title_cell.font = Font(bold=True, size=13, color="1A3C6E")
        title_cell.alignment = center
        ws.row_dimensions[1].height = 30

        # Заголовки колонок
        headers = ['#', 'Трек-код', 'Клиент (ID)', 'Описание', 'Статус', 'Ячейка', 'Вес (кг)', 'К оплате ($)', 'Дата добавления']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        # Ширины колонок
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 18
        ws.column_dimensions['C'].width = 16
        ws.column_dimensions['D'].width = 28
        ws.column_dimensions['E'].width = 22
        ws.column_dimensions['F'].width = 10
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 14
        ws.column_dimensions['I'].width = 18

        STATUS_MAP = dict(Package.STATUS_CHOICES)
        packages = Package.objects.filter(
            created_at__gte=date_from, created_at__lte=date_to
        ).select_related('client', 'client__user').order_by('-created_at')

        row_fill_even = PatternFill("solid", fgColor="F0F4FF")
        total_price = 0
        total_weight = 0

        for i, pkg in enumerate(packages, start=1):
            row = i + 2
            fill = row_fill_even if i % 2 == 0 else None
            data = [
                i,
                pkg.track_code,
                pkg.client.client_code if pkg.client else '—',
                pkg.description or '—',
                STATUS_MAP.get(pkg.status, pkg.status),
                pkg.shelf_location or '—',
                float(pkg.weight),
                float(pkg.total_price),
                pkg.created_at.strftime('%d.%m.%Y %H:%M'),
            ]
            for col, val in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = border
                cell.alignment = center
                if fill:
                    cell.fill = fill

            total_price += float(pkg.total_price)
            total_weight += float(pkg.weight)

        # Итоговая строка
        total_row = packages.count() + 3
        summary_fill = PatternFill("solid", fgColor="E8F0FE")
        ws.cell(row=total_row, column=1, value="ИТОГО").font = Font(bold=True)
        ws.cell(row=total_row, column=7, value=round(total_weight, 2)).font = Font(bold=True)
        ws.cell(row=total_row, column=8, value=round(total_price, 2)).font = Font(bold=True)
        for col in range(1, 10):
            ws.cell(row=total_row, column=col).fill = summary_fill
            ws.cell(row=total_row, column=col).border = border

    # ─────────────────────────────────────────
    #  ОТЧЁТ 2: КЛИЕНТЫ
    # ─────────────────────────────────────────
    def _build_clients_sheet(self, wb, date_from, date_to):
        ws = wb.active
        ws.title = "Клиенты"

        header_fill = PatternFill("solid", fgColor="1A3C6E")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="D0D7E3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.merge_cells('A1:F1')
        ws['A1'].value = f"Отчет по клиентам — Kayhon Cargo | Дата: {datetime.now().strftime('%d.%m.%Y')}"
        ws['A1'].font = Font(bold=True, size=13, color="1A3C6E")
        ws['A1'].alignment = center

        headers = ['#', 'ID Клиента', 'Имя', 'Телефон', 'Telegram ID', 'Кол-во посылок']
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        col_widths = [5, 18, 22, 18, 18, 16]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        clients = ClientProfile.objects.filter(role='client').select_related('user').order_by('-user__date_joined')
        row_fill_even = PatternFill("solid", fgColor="F0F4FF")

        for i, client in enumerate(clients, start=1):
            row = i + 2
            fill = row_fill_even if i % 2 == 0 else None
            data = [
                i,
                client.client_code,
                client.user.first_name or '—',
                client.phone_number,
                client.telegram_id or '—',
                client.packages.count(),
            ]
            for col, val in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = border
                cell.alignment = center
                if fill:
                    cell.fill = fill

    # ─────────────────────────────────────────
    #  ОТЧЁТ 3: ФИНАНСЫ
    # ─────────────────────────────────────────
    def _build_finance_sheet(self, wb, date_from, date_to):
        ws = wb.active
        ws.title = "Финансы"

        header_fill = PatternFill("solid", fgColor="14532d")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        center = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="D0D7E3")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        period_str = f"{date_from.strftime('%d.%m.%Y')} – {date_to.strftime('%d.%m.%Y')}"
        ws.merge_cells('A1:G1')
        ws['A1'].value = f"Финансовый отчет — Kayhon Cargo | {period_str}"
        ws['A1'].font = Font(bold=True, size=13, color="14532d")
        ws['A1'].alignment = center

        headers = ['#', 'Трек-код', 'Клиент (ID)', 'Вес (кг)', 'К оплате ($)', 'Оплачено', 'Дата выдачи']
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        col_widths = [5, 18, 16, 12, 14, 12, 20]
        for i, w in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

        packages = Package.objects.filter(
            created_at__gte=date_from,
            created_at__lte=date_to,
            status='delivered'
        ).select_related('client').order_by('-payment_date')

        row_fill_even = PatternFill("solid", fgColor="F0FFF4")
        total_price = 0
        total_weight = 0

        for i, pkg in enumerate(packages, start=1):
            row = i + 2
            fill = row_fill_even if i % 2 == 0 else None
            data = [
                i,
                pkg.track_code,
                pkg.client.client_code if pkg.client else '—',
                float(pkg.weight),
                float(pkg.total_price),
                'Да' if pkg.is_paid else 'Нет',
                pkg.payment_date.strftime('%d.%m.%Y') if pkg.payment_date else '—',
            ]
            for col, val in enumerate(data, start=1):
                cell = ws.cell(row=row, column=col, value=val)
                cell.border = border
                cell.alignment = center
                if fill:
                    cell.fill = fill
            total_price += float(pkg.total_price)
            total_weight += float(pkg.weight)

        # Итог
        summary_row = packages.count() + 3
        summary_fill = PatternFill("solid", fgColor="DCFCE7")
        ws.cell(row=summary_row, column=1, value="ИТОГО").font = Font(bold=True)
        ws.cell(row=summary_row, column=4, value=round(total_weight, 2)).font = Font(bold=True)
        ws.cell(row=summary_row, column=5, value=round(total_price, 2)).font = Font(bold=True)
        for col in range(1, 8):
            ws.cell(row=summary_row, column=col).fill = summary_fill
            ws.cell(row=summary_row, column=col).border = border
